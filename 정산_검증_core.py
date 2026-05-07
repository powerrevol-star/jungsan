"""
정산 검증 Core — 내부망 전용
==============================
pandas / numpy 의존성 없이 openpyxl + 표준 라이브러리만 사용.

플랫폼(내부 LLM 에이전트)에서 아래 형식의 JSON을 받아 run_from_json()을 호출.

입력 JSON 예시:
    {
        "affiliate": "KB저축은행",
        "settlement_month": "2026-05"   # 또는 "05" (연도 없는 경우)
    }

반환값: dict (결과 딕셔너리, JSON 직렬화 가능)
"""

import json
import math
import os
import re
import sys
from datetime import datetime
from openpyxl import load_workbook


# ─────────────────────────────────────────────
# 0. 경로 설정
# ─────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# DB 총본 파일명 (고정)
DB_FILENAME = "2604_F_기반_Agent_자료.xlsx"
DB_SHEET    = "대출시트"


# ─────────────────────────────────────────────
# 1. 설정 상수
# ─────────────────────────────────────────────

# DB 컬럼명(헤더) → 표준명 매핑
DB_COL_MAP = {
    "MYDA_ORG_NM":       "제휴사명",
    "SINCDT":            "신청일자",
    "DC_EXE_DT":         "실행일자",
    "DC_SINC_NO":        "대출신청번호",
    "JEHYUSA_DC_PRDT_C": "대출상품코드",
    "DC_PRDT_NM":        "대출상품명",
    "DCAMT":             "대출금액",
    "결과":              "상태",
    "수수료":            "지급수수료",
}

# 제휴사 회신 파일 컬럼명 → 표준명 매핑 (공백 포함 변형 모두 처리)
PARTNER_COL_MAP = {
    "신청일자":      "신청일자",
    "실행일자":      "실행일자",
    "대출 신청번호": "대출신청번호",
    "대출신청번호":  "대출신청번호",
    "대출 상품코드": "대출상품코드",
    "대출상품코드":  "대출상품코드",
    "대출 상품명":   "대출상품명",
    "대출상품명":    "대출상품명",
    "대출금액":      "대출금액",
    "상태":          "상태",
    "지급수수료":    "지급수수료",
}

# 검증 대상 컬럼 (표준명 기준)
CHECK_COLS = ["신청일자", "실행일자", "대출상품코드", "대출상품명", "대출금액", "상태", "지급수수료"]

# 상태값 동의어 그룹: (대표값, frozenset)
STATUS_SYNONYM_GROUPS = [
    ("실행",  frozenset({"실행", "정상", "03-기표(정상)", "기표"})),
    ("철회",  frozenset({"철회", "취소", "해지", "06-철회"})),
    ("심사중", frozenset({"심사중", "처리중"})),
    ("완제",  frozenset({"완제", "종료", "05-종료(완제)"})),
]


# ─────────────────────────────────────────────
# 2. 유틸리티
# ─────────────────────────────────────────────

def _is_nan(v) -> bool:
    """None 또는 float NaN 판별."""
    if v is None:
        return True
    try:
        return math.isnan(float(v))
    except (TypeError, ValueError):
        return False


def _to_float(v):
    """숫자 변환 실패 시 None 반환."""
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _fmt_amount(v) -> str:
    """금액을 '1,234,567원' 형식으로 포맷."""
    f = _to_float(v)
    if f is None or _is_nan(f):
        return "-"
    return f"{int(f):,}원"


def _normalize_date(v) -> str:
    """날짜를 YYYYMMDD 8자리 문자열로 정규화."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d")
    s = str(v).strip().replace("-", "").replace("/", "").replace(".", "")
    if "." in s:          # float 형태(20260407.0)
        s = s.split(".")[0]
    return s[:8] if len(s) >= 8 else s


def _normalize_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_status(v: str) -> str:
    """상태값 → 대표값 정규화 (2단계: 정확한 일치 → 부분 문자열)."""
    s = str(v).strip()
    for rep, synonyms in STATUS_SYNONYM_GROUPS:
        if s in synonyms:
            return rep
    for rep, synonyms in STATUS_SYNONYM_GROUPS:
        for kw in synonyms:
            if kw in s:
                return rep
    return s


def _normalize_month(raw: str) -> str:
    """
    settlement_month 정규화 → 'YYYY-MM' 문자열.
    입력: '2026-05', '05', '4', '2026/05', '202605' 등
    """
    raw = str(raw).strip()
    # YYYY-MM or YYYY/MM
    m = re.match(r"^(\d{4})[-/](\d{1,2})$", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    # YYYYMM
    m = re.match(r"^(\d{4})(\d{2})$", raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # MM or M only → 현재 연도 사용
    m = re.match(r"^(\d{1,2})$", raw)
    if m:
        year = datetime.today().year
        return f"{year}-{int(m.group(1)):02d}"
    return raw  # 파싱 불가 시 원본 반환


# ─────────────────────────────────────────────
# 3. Excel 읽기 (openpyxl)
# ─────────────────────────────────────────────

def _read_sheet(path: str, sheet_name: str) -> list[dict]:
    """
    xlsx 시트를 읽어 [{컬럼명: 값, ...}, ...] 리스트로 반환.
    첫 행이 헤더라고 가정.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else f"_col{j}"
                       for j, c in enumerate(row)]
        else:
            rows.append(dict(zip(headers, row)))

    wb.close()
    return rows


def _find_header_row(path: str, keyword: str = "신청일자", max_rows: int = 10) -> int:
    """헤더 행 인덱스 자동 탐지 (keyword가 포함된 첫 번째 행)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        if any(keyword in str(c) for c in row if c is not None):
            result = i
            break
    wb.close()
    return result


def _read_partner_sheet(path: str) -> list[dict]:
    """
    제휴사 회신 파일 읽기.
    헤더 위치를 자동 탐지하고, 대출신청번호가 유효한 행만 반환.
    """
    header_row_idx = _find_header_row(path, keyword="신청일자")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row_idx:
            continue
        if i == header_row_idx:
            headers = [str(c).strip() if c is not None else f"_col{j}"
                       for j, c in enumerate(row)]
            continue
        if not headers:
            continue
        record = dict(zip(headers, row))
        rows.append(record)

    wb.close()

    # 컬럼명 표준화
    normalized = []
    for rec in rows:
        new_rec = {}
        for k, v in rec.items():
            std_key = PARTNER_COL_MAP.get(k, k)
            new_rec[std_key] = v
        normalized.append(new_rec)

    return normalized


# ─────────────────────────────────────────────
# 4. 데이터 로드
# ─────────────────────────────────────────────

def load_db_rows(db_path: str, affiliate: str) -> list[dict]:
    """
    DB 총본에서 특정 제휴사 행만 필터링하여 표준 컬럼명으로 반환.
    정확한 이름 일치 우선 → 없으면 부분 일치.
    """
    all_rows = _read_sheet(db_path, DB_SHEET)

    # 컬럼명 표준화
    std_rows = []
    for rec in all_rows:
        new_rec = {DB_COL_MAP.get(k, k): v for k, v in rec.items()}
        std_rows.append(new_rec)

    # 정확한 이름 일치 우선
    exact = [r for r in std_rows if _normalize_str(r.get("제휴사명")) == affiliate]
    if exact:
        return exact

    # 부분 일치 fallback
    escaped = re.escape(affiliate)
    partial = [r for r in std_rows
               if re.search(escaped, _normalize_str(r.get("제휴사명", "")))]
    if partial:
        return partial

    all_names = sorted({_normalize_str(r.get("제휴사명", "")) for r in std_rows if r.get("제휴사명")})
    raise ValueError(f"DB에서 '{affiliate}' 데이터를 찾을 수 없습니다. 존재하는 제휴사: {all_names}")


def load_partner_rows(partner_path: str, db_key_len: int = 14) -> list[dict]:
    """
    제휴사 회신 파일 로드.
    대출신청번호가 유효한(영숫자 10자 이상) 행만 유지하고,
    DB 키 길이만큼만 신청번호를 잘라냄(상품코드 suffix 제거).
    """
    rows = _read_partner_sheet(partner_path)

    result = []
    for rec in rows:
        loan_no = _normalize_str(rec.get("대출신청번호", ""))
        # 유효한 신청번호: 영숫자 10자 이상
        if not re.match(r"^[A-Za-z0-9]{10,}", loan_no):
            continue
        rec["대출신청번호"] = loan_no[:db_key_len]
        result.append(rec)

    return result


def get_db_partner_names(db_path: str) -> list[str]:
    """DB 파일에서 제휴사명 목록 추출."""
    rows = _read_sheet(db_path, DB_SHEET)
    names = set()
    for r in rows:
        v = r.get("MYDA_ORG_NM")
        if v:
            names.add(str(v).strip())
    return sorted(names)


# ─────────────────────────────────────────────
# 5. 데이터 정규화
# ─────────────────────────────────────────────

def _normalize_row(rec: dict) -> dict:
    """단일 행 정규화."""
    out = dict(rec)
    date_cols   = {"신청일자", "실행일자"}
    amount_cols = {"대출금액", "지급수수료"}
    str_cols    = {"대출신청번호", "대출상품코드", "대출상품명", "상태"}

    for k in list(out.keys()):
        if k in date_cols:
            out[k] = _normalize_date(out[k])
        elif k in amount_cols:
            out[k] = _to_float(out[k])
        elif k in str_cols:
            out[k] = _normalize_str(out[k])
    return out


def _normalize_rows(rows: list[dict]) -> list[dict]:
    return [_normalize_row(r) for r in rows]


# ─────────────────────────────────────────────
# 6. 중복 신청번호 합산
# ─────────────────────────────────────────────

def _group_by_loan_no(rows: list[dict]) -> dict[str, dict]:
    """
    대출신청번호를 키로 그룹화.
    중복 시 대출금액·지급수수료는 합산, 나머지는 첫 번째 행 값 사용.
    """
    grouped: dict[str, dict] = {}
    AMOUNT_COLS = {"대출금액", "지급수수료"}

    for rec in rows:
        key = rec.get("대출신청번호", "")
        if not key:
            continue
        if key not in grouped:
            grouped[key] = dict(rec)
            grouped[key]["_count"] = 1
        else:
            grouped[key]["_count"] += 1
            for col in AMOUNT_COLS:
                prev = grouped[key].get(col)
                curr = rec.get(col)
                if not _is_nan(curr):
                    grouped[key][col] = (prev or 0) + curr
    return grouped


# ─────────────────────────────────────────────
# 7. 비교 로직
# ─────────────────────────────────────────────

def _compare_value(col: str, db_val, pt_val) -> bool:
    if col in ("대출금액", "지급수수료"):
        db_f = _to_float(db_val)
        pt_f = _to_float(pt_val)
        if _is_nan(db_f) and _is_nan(pt_f):
            return True
        if _is_nan(db_f) or _is_nan(pt_f):
            return False
        return abs(db_f - pt_f) < 1  # 1원 미만 오차 허용
    elif col == "상태":
        return _normalize_status(str(db_val)) == _normalize_status(str(pt_val))
    else:
        return str(db_val).strip() == str(pt_val).strip()


# ─────────────────────────────────────────────
# 8. 검증 핵심 로직
# ─────────────────────────────────────────────

def run_verification(db_rows: list[dict], partner_rows: list[dict]) -> dict:
    """
    대출신청번호 기준으로 DB vs 제휴사 회신자료 검증.
    중복 신청번호는 금액 합산 후 비교.
    """
    db_rows     = _normalize_rows(db_rows)
    partner_rows = _normalize_rows(partner_rows)

    db_map  = _group_by_loan_no(db_rows)
    pt_map  = _group_by_loan_no(partner_rows)

    all_keys = sorted(set(db_map.keys()) | set(pt_map.keys()))

    # 중복 신청번호 정보 수집
    duplicate_keys = []
    for key in all_keys:
        db_cnt = db_map[key].get("_count", 1) if key in db_map else 0
        pt_cnt = pt_map[key].get("_count", 1) if key in pt_map else 0
        if db_cnt > 1 or pt_cnt > 1:
            duplicate_keys.append({
                "대출신청번호":    key,
                "DB건수":         db_cnt,
                "제휴사건수":      pt_cnt,
                "DB_합산대출금액":  _fmt_amount(db_map[key].get("대출금액")) if key in db_map else "-",
                "제휴_합산대출금액": _fmt_amount(pt_map[key].get("대출금액")) if key in pt_map else "-",
                "DB_합산지급수수료":  _fmt_amount(db_map[key].get("지급수수료")) if key in db_map else "-",
                "제휴_합산지급수수료": _fmt_amount(pt_map[key].get("지급수수료")) if key in pt_map else "-",
                "비고": "금액 합산 후 비교",
            })

    matched         = []
    mismatch_records = []
    only_in_db      = []
    only_in_partner  = []

    for key in all_keys:
        in_db  = key in db_map
        in_pt  = key in pt_map

        if in_db and not in_pt:
            r = db_map[key]
            only_in_db.append({
                "대출신청번호":  key,
                "DB_신청일자":   r.get("신청일자", ""),
                "DB_대출상품명": r.get("대출상품명", ""),
                "DB_대출금액":   _fmt_amount(r.get("대출금액")),
                "DB_상태":       r.get("상태", ""),
                "DB_지급수수료": _fmt_amount(r.get("지급수수료")),
            })
            continue

        if in_pt and not in_db:
            r = pt_map[key]
            only_in_partner.append({
                "대출신청번호":      key,
                "제휴사_신청일자":   r.get("신청일자", ""),
                "제휴사_대출상품명": r.get("대출상품명", ""),
                "제휴사_대출금액":   _fmt_amount(r.get("대출금액")),
                "제휴사_상태":       r.get("상태", ""),
                "제휴사_지급수수료": _fmt_amount(r.get("지급수수료")),
            })
            continue

        # 양쪽 모두 존재 → 항목별 비교
        db_r = db_map[key]
        pt_r = pt_map[key]

        # 상품코드 동일하면 상품명 비교 면제
        same_code = (
            _normalize_str(db_r.get("대출상품코드", "")) ==
            _normalize_str(pt_r.get("대출상품코드", "")) and
            _normalize_str(db_r.get("대출상품코드", "")) != ""
        )

        col_mismatches = []
        for col in CHECK_COLS:
            if col == "대출상품명" and same_code:
                continue
            db_val = db_r.get(col, "")
            pt_val = pt_r.get(col, "")
            if not _compare_value(col, db_val, pt_val):
                col_mismatches.append({
                    "항목":   col,
                    "DB값":   _fmt_amount(db_val) if col in ("대출금액", "지급수수료") else str(db_val),
                    "제휴사값": _fmt_amount(pt_val) if col in ("대출금액", "지급수수료") else str(pt_val),
                })

        if col_mismatches:
            mismatch_records.append({
                "대출신청번호": key,
                "불일치항목":  col_mismatches,
            })
        else:
            matched.append(key)

    return {
        "db_total":         len(db_map),
        "partner_total":    len(pt_map),
        "matched":          matched,
        "mismatch_records": mismatch_records,
        "only_in_db":       only_in_db,
        "only_in_partner":  only_in_partner,
        "duplicate_keys":   duplicate_keys,
    }


# ─────────────────────────────────────────────
# 9. 진입점: JSON 입력 → 검증 실행
# ─────────────────────────────────────────────

def run_from_json(input_json: dict | str, partner_file_path: str) -> dict:
    """
    플랫폼(내부 LLM 에이전트)에서 호출하는 진입점.

    Parameters
    ----------
    input_json : dict 또는 JSON 문자열
        {"affiliate": "KB저축은행", "settlement_month": "2026-05"}
        settlement_month는 "YYYY-MM" 또는 "MM" 형식 모두 허용.

    partner_file_path : str
        제휴사 회신 Excel 파일 경로 (절대경로 권장).

    Returns
    -------
    dict
        {
          "affiliate":         str,   # 제휴사명
          "settlement_month":  str,   # 정규화된 정산월 (YYYY-MM)
          "db_total":          int,
          "partner_total":     int,
          "matched_count":     int,
          "mismatch_count":    int,
          "only_in_db_count":  int,
          "only_in_partner_count": int,
          "duplicate_count":   int,
          "mismatch_records":  list,
          "only_in_db":        list,
          "only_in_partner":   list,
          "duplicate_keys":    list,
          "error":             str | None,
        }
    """
    # ── 입력 파싱 ──
    if isinstance(input_json, str):
        input_json = json.loads(input_json)

    affiliate        = str(input_json.get("affiliate", "")).strip()
    settlement_month = _normalize_month(str(input_json.get("settlement_month", "")))

    if not affiliate:
        return {"error": "affiliate 값이 비어 있습니다.", "affiliate": "", "settlement_month": settlement_month}

    # ── 파일 경로 확인 ──
    db_path = os.path.join(BASE_DIR, DB_FILENAME)
    if not os.path.exists(db_path):
        return {"error": f"DB 파일을 찾을 수 없습니다: {db_path}", "affiliate": affiliate, "settlement_month": settlement_month}
    if not os.path.exists(partner_file_path):
        return {"error": f"제휴사 파일을 찾을 수 없습니다: {partner_file_path}", "affiliate": affiliate, "settlement_month": settlement_month}

    try:
        # ── DB 로드 ──
        db_rows = load_db_rows(db_path, affiliate)

        # DB 신청번호 길이 자동 파악
        key_lengths = [len(r.get("대출신청번호", "") or "") for r in db_rows if r.get("대출신청번호")]
        db_key_len = max(set(key_lengths), key=key_lengths.count) if key_lengths else 14

        # ── 제휴사 로드 ──
        partner_rows = load_partner_rows(partner_file_path, db_key_len=db_key_len)

        # ── 검증 실행 ──
        result = run_verification(db_rows, partner_rows)

        return {
            "affiliate":              affiliate,
            "settlement_month":       settlement_month,
            "db_total":               result["db_total"],
            "partner_total":          result["partner_total"],
            "matched_count":          len(result["matched"]),
            "mismatch_count":         len(result["mismatch_records"]),
            "only_in_db_count":       len(result["only_in_db"]),
            "only_in_partner_count":  len(result["only_in_partner"]),
            "duplicate_count":        len(result["duplicate_keys"]),
            "mismatch_records":       result["mismatch_records"],
            "only_in_db":             result["only_in_db"],
            "only_in_partner":        result["only_in_partner"],
            "duplicate_keys":         result["duplicate_keys"],
            "error":                  None,
        }

    except ValueError as e:
        return {"error": str(e), "affiliate": affiliate, "settlement_month": settlement_month}
    except Exception as e:
        return {"error": f"검증 중 오류 발생: {e}", "affiliate": affiliate, "settlement_month": settlement_month}


# ─────────────────────────────────────────────
# 10. CLI 실행 (테스트용)
# ─────────────────────────────────────────────

def _print_report(result: dict):
    SEP = "=" * 65
    SUB = "-" * 65
    print(SEP)
    print(f"  [{result['affiliate']}] 대출 정산 검증 보고서  ({result['settlement_month']})")
    print(f"  기준: {datetime.today().strftime('%Y-%m-%d')}")
    print(SEP)
    if result.get("error"):
        print(f"\n  [오류] {result['error']}\n")
        return

    print(f"\n{'[ 요약 ]'}")
    print(SUB)
    print(f"  제휴사 회신자료 기준 총 건수  : {result['partner_total']:>4}건")
    print(f"  DB 자료 기준 총 건수          : {result['db_total']:>4}건")
    print(f"  완전 일치                     : {result['matched_count']:>4}건")
    total_mis = result['mismatch_count'] + result['only_in_db_count'] + result['only_in_partner_count']
    print(f"  불일치 (합계)                 : {total_mis:>4}건")
    print(f"    ├ 값 불일치 (양쪽 존재)     : {result['mismatch_count']:>4}건")
    print(f"    ├ DB에만 존재               : {result['only_in_db_count']:>4}건")
    print(f"    └ 제휴사 회신에만 존재      : {result['only_in_partner_count']:>4}건")
    if result['duplicate_count'] > 0:
        print(f"\n  ⚠ 중복 신청번호             : {result['duplicate_count']:>4}건 (금액 합산 후 비교)")
    print(SUB)

    if result["mismatch_records"]:
        print(f"\n【 값 불일치 명세 ({result['mismatch_count']}건) 】")
        for i, rec in enumerate(result["mismatch_records"], 1):
            print(f"\n  [{i}] 대출신청번호: {rec['대출신청번호']}")
            for mis in rec["불일치항목"]:
                print(f"      ▸ {mis['항목']}")
                print(f"        DB     : {mis['DB값']}")
                print(f"        제휴사 : {mis['제휴사값']}")

    if result["only_in_db"]:
        print(f"\n【 DB에만 존재 ({result['only_in_db_count']}건) 】")
        for i, rec in enumerate(result["only_in_db"], 1):
            print(f"\n  [{i}] {rec['대출신청번호']}  {rec['DB_대출상품명']}  {rec['DB_대출금액']}  {rec['DB_상태']}")

    if result["only_in_partner"]:
        print(f"\n【 제휴사에만 존재 ({result['only_in_partner_count']}건) 】")
        for i, rec in enumerate(result["only_in_partner"], 1):
            print(f"\n  [{i}] {rec['대출신청번호']}  {rec['제휴사_대출상품명']}  {rec['제휴사_대출금액']}  {rec['제휴사_상태']}")

    print(f"\n{SEP}\n")


if __name__ == "__main__":
    """
    CLI 테스트:
        python3 정산_검증_core.py '{"affiliate":"BNK저축은행","settlement_month":"2026-04"}' BNK저축은행_2604_대출.xlsx
        python3 정산_검증_core.py '{"affiliate":"KB저축은행","settlement_month":"04"}' KB저축은행_26년4월_대출정산.xlsx
    """
    if len(sys.argv) < 3:
        print("사용법: python3 정산_검증_core.py '<JSON>' <제휴사파일.xlsx>")
        sys.exit(1)

    input_arg    = sys.argv[1]
    partner_file = sys.argv[2]
    if not os.path.isabs(partner_file):
        partner_file = os.path.join(BASE_DIR, partner_file)

    result = run_from_json(input_arg, partner_file)
    _print_report(result)
